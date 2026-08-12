#!/usr/bin/env python3
"""Test bout-en-bout: extraction de toutes les factures + export Excel."""

from __future__ import annotations

import asyncio
import sys
from datetime import date
from io import BytesIO
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).parent.parent))

from excel_export import export_to_bytes
from invoice_extractor import extract_invoice, merge_extractions
from models import ExportRequest

INVOICES_DIR = Path(__file__).parent / "invoices"


async def main() -> int:
    pdfs = sorted(INVOICES_DIR.glob("*.pdf"))
    results = []
    for pdf in pdfs:
        result = await extract_invoice(pdf.name, pdf.read_bytes(), "application/pdf")
        results.append(result)
        print(f"{pdf.name}: {len(result.lines)} ligne(s)")

    lines = merge_extractions(results)
    print(f"\nTotal lignes exportables: {len(lines)}")

    request = ExportRequest(
        client_name="Aichoum",
        period="062026",
        lines=lines,
    )
    content = export_to_bytes(request)
    out = Path(__file__).parent / "output_test.xlsx"
    out.write_bytes(content)
    print(f"Fichier généré: {out}")

    wb = openpyxl.load_workbook(BytesIO(content), data_only=True)
    ws = wb["EDI0626"]
    print(f"Lignes dans Excel: {ws.max_row - 1}")
    for row in range(2, ws.max_row + 1):
        fact = ws.cell(row, 2).value
        frss = ws.cell(row, 8).value
        ht = ws.cell(row, 4).value
        code = ws.cell(row, 14).value
        print(f"  - {fact} | {frss} | HT={ht} | CODE TVA={code}")

    return 0


if __name__ == "__main__":
    raise SystemExit(asyncio.run(main()))
