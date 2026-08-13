from __future__ import annotations

from datetime import date, datetime
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from models import ExportRequest, InvoiceLine

HEADERS = [
    "OR",
    "FACT_NUM",
    "DESIGNATION",
    "M_HT",
    "TVA",
    "M_TTC",
    "IF",
    "LIB_FRSS",
    "ICE_FRS",
    "TAUX",
    "ID_PAIE",
    "DATE_PAIE",
    "DATE_FAC",
    "CODE TVA",
]

COLUMN_WIDTHS = [5, 14, 22, 11, 10, 11, 10, 20, 18, 7, 8, 12, 12, 10]

HEADER_FILL = PatternFill("solid", fgColor="0B6BCB")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
THIN_BORDER = Border(
    left=Side(style="thin", color="D9E3EF"),
    right=Side(style="thin", color="D9E3EF"),
    top=Side(style="thin", color="D9E3EF"),
    bottom=Side(style="thin", color="D9E3EF"),
)
BODY_BORDER = Border(
    left=Side(style="thin", color="E8EEF4"),
    right=Side(style="thin", color="E8EEF4"),
    top=Side(style="thin", color="E8EEF4"),
    bottom=Side(style="thin", color="E8EEF4"),
)

DATE_FMT = "dd/mm/yyyy"
AMOUNT_FMT = "#,##0.00"
TAUX_FMT = "0%"

TEMPLATE_PATH = Path(__file__).parent / "templates" / "ded_tva_template.xlsx"


def _to_excel_date(value: date | datetime | None) -> datetime | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.replace(hour=0, minute=0, second=0, microsecond=0)
    return datetime(value.year, value.month, value.day)


def _write_line(ws, row_idx: int, line: InvoiceLine) -> None:
    values = [
        line.or_value,
        line.fact_num,
        line.designation.value,
        line.m_ht,
        line.tva,
        line.m_ttc,
        line.if_fournisseur,
        line.lib_frss,
        line.ice_frs,
        line.taux,
        line.id_paie,
        _to_excel_date(line.date_paie),
        _to_excel_date(line.date_fac),
        line.resolved_code_tva(),
    ]
    amount_cols = {4, 5, 6}
    date_cols = {12, 13}
    for col_idx, value in enumerate(values, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.border = BODY_BORDER
        if col_idx in amount_cols:
            cell.number_format = AMOUNT_FMT
            cell.alignment = Alignment(horizontal="right")
        elif col_idx == 10:
            cell.number_format = TAUX_FMT
            cell.alignment = Alignment(horizontal="center")
        elif col_idx in date_cols and value is not None:
            cell.number_format = DATE_FMT
            cell.alignment = Alignment(horizontal="center")
        elif col_idx in {1, 11, 14}:
            cell.alignment = Alignment(horizontal="center")


def _ensure_headers(ws) -> None:
    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def _apply_sheet_layout(ws) -> None:
    for col_idx, width in enumerate(COLUMN_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = "A2"
    if ws.max_row > 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{ws.max_row}"


def build_workbook(request: ExportRequest) -> Workbook:
    # Période MMAAAA → feuille EDIMMAA (ex: 062026 → EDI0626)
    sheet_name = request.sheet_name or f"EDI{request.period[:2]}{request.period[4:6]}"

    if TEMPLATE_PATH.exists():
        wb = load_workbook(TEMPLATE_PATH)
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            if ws.max_row > 1:
                ws.delete_rows(2, ws.max_row)
        else:
            ws = wb.create_sheet(sheet_name)
            _ensure_headers(ws)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        _ensure_headers(ws)

    for idx, line in enumerate(request.lines, start=2):
        _write_line(ws, idx, line)

    _apply_sheet_layout(ws)
    return wb


def export_to_bytes(request: ExportRequest) -> bytes:
    wb = build_workbook(request)
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def export_filename(client_name: str, period: str) -> str:
    safe_client = "".join(ch if ch.isalnum() or ch in "-_" else "_" for ch in client_name.strip())
    return f"{safe_client}_DED_TVA_{period}.xlsx"
