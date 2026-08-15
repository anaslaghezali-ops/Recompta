"""Parse CSV/Excel bank statements (same rules as bank-statement-client.js)."""

from __future__ import annotations

import csv
import io
import re
import unicodedata
from typing import Any

from openpyxl import load_workbook

FEE_KEYWORDS = re.compile(
    r"commission|frais\s*banc|agios|tenue\s*de\s*compte|cotisation|carte\s*banc|"
    r"retrait\s*dab|frais\s*de\s*|pakage|package|interet|intérêt",
    re.IGNORECASE,
)
SKIP_KEYWORDS = re.compile(
    r"virement\s+recu|virement\s+reçu|remise\s+cheque|remise\s+chèque|"
    r"depot\s+especes|dépôt\s+espèces|solde\s+initial|solde\s+final|total\s+mouvement",
    re.IGNORECASE,
)
DATE_HEADER = re.compile(r"date|valeur|opération|operation", re.IGNORECASE)
LABEL_HEADER = re.compile(r"libell|description|intitul|motif|détail|detail", re.IGNORECASE)
DEBIT_HEADER = re.compile(r"débit|debit|montant\s*débit", re.IGNORECASE)
CREDIT_HEADER = re.compile(r"crédit|credit|montant\s*crédit", re.IGNORECASE)
AMOUNT_HEADER = re.compile(r"^montant$|amount|somme", re.IGNORECASE)


def normalize_text(value: Any) -> str:
    text = unicodedata.normalize("NFD", str(value or ""))
    text = "".join(ch for ch in text if unicodedata.category(ch) != "Mn")
    return re.sub(r"\s+", " ", text.upper()).strip()


def parse_amount(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(" ", "")
    if not text:
        return None
    negative = text.startswith("-") or text.startswith("(")
    text = re.sub(r"[()]", "", text)
    text = re.sub(r"[^\d,.-]", "", text)
    if "," in text and "." in text:
        if text.rfind(",") > text.rfind("."):
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    elif "," in text:
        text = text.replace(",", ".")
    try:
        num = float(text)
    except ValueError:
        return None
    return -abs(num) if negative else num


def parse_date(value: Any) -> str:
    if value is None or value == "":
        return ""
    if isinstance(value, (int, float)) and 30000 < value < 60000:
        from datetime import date, timedelta

        epoch = date(1899, 12, 30)
        parsed = epoch + timedelta(days=int(value))
        return parsed.isoformat()
    text = str(value).strip()
    iso = re.match(r"^(\d{4})-(\d{2})-(\d{2})", text)
    if iso:
        return f"{iso.group(1)}-{iso.group(2)}-{iso.group(3)}"
    fr = re.match(r"^(\d{1,2})[/.-](\d{1,2})[/.-](\d{2,4})", text)
    if fr:
        year = fr.group(3)
        if len(year) == 2:
            year = f"20{year}"
        return f"{year}-{fr.group(2).zfill(2)}-{fr.group(1).zfill(2)}"
    return ""


def detect_delimiter(text: str) -> str:
    first_line = text.splitlines()[0] if text.splitlines() else ""
    semicolons = first_line.count(";")
    commas = first_line.count(",")
    return ";" if semicolons >= commas else ","


def find_header_row(rows: list[list[Any]]) -> int:
    for index, row in enumerate(rows[:15]):
        if not row:
            continue
        joined = " ".join(normalize_text(cell) for cell in row)
        if DATE_HEADER.search(joined) and (
            LABEL_HEADER.search(joined) or DEBIT_HEADER.search(joined) or AMOUNT_HEADER.search(joined)
        ):
            return index
    return 0


def map_columns(header_row: list[Any]) -> dict[str, int]:
    cols = {"date": -1, "label": -1, "debit": -1, "credit": -1, "amount": -1}
    for index, cell in enumerate(header_row):
        header = normalize_text(cell)
        if DATE_HEADER.search(header) and cols["date"] < 0:
            cols["date"] = index
        elif LABEL_HEADER.search(header) and cols["label"] < 0:
            cols["label"] = index
        elif DEBIT_HEADER.search(header) and cols["debit"] < 0:
            cols["debit"] = index
        elif CREDIT_HEADER.search(header) and cols["credit"] < 0:
            cols["credit"] = index
        elif AMOUNT_HEADER.search(header) and cols["amount"] < 0:
            cols["amount"] = index
    if cols["label"] < 0:
        for index, cell in enumerate(header_row):
            if len(str(cell).strip()) > 2:
                cols["label"] = index
                break
    return cols


def row_to_transaction(row: list[Any], cols: dict[str, int], index: int) -> dict[str, Any] | None:
    label = str(row[cols["label"]] if cols["label"] >= 0 and cols["label"] < len(row) else "").strip()
    date_col = cols["date"]
    date = parse_date(row[date_col] if date_col >= 0 and date_col < len(row) else "")
    amount: float | None = None

    if cols["debit"] >= 0 or cols["credit"] >= 0:
        debit = parse_amount(row[cols["debit"]]) if cols["debit"] >= 0 and cols["debit"] < len(row) else None
        credit = parse_amount(row[cols["credit"]]) if cols["credit"] >= 0 and cols["credit"] < len(row) else None
        if debit and abs(debit) > 0:
            amount = -abs(debit)
        elif credit and abs(credit) > 0:
            amount = abs(credit)

    if amount is None and cols["amount"] >= 0 and cols["amount"] < len(row):
        amount = parse_amount(row[cols["amount"]])

    if not date or amount is None or abs(amount) < 0.01:
        return None

    normalized_label = normalize_text(label)
    if SKIP_KEYWORDS.search(normalized_label):
        return None

    is_fee = bool(FEE_KEYWORDS.search(normalized_label))
    is_debit = amount < 0
    if is_fee and is_debit:
        txn_type = "fee"
    elif is_debit:
        txn_type = "payment"
    else:
        txn_type = "credit"

    abs_amount = abs(amount)
    return {
        "id": f"row-{index}",
        "date": date,
        "label": label,
        "amount": abs_amount * (-1 if is_debit else 1),
        "absAmount": abs_amount,
        "type": txn_type,
    }


def parse_bank_rows(rows: list[list[Any]]) -> tuple[list[dict[str, Any]], list[str]]:
    if not rows:
        return [], ["Fichier vide ou illisible."]

    header_index = find_header_row(rows)
    cols = map_columns(rows[header_index] if header_index < len(rows) else [])
    warnings: list[str] = []

    if cols["date"] < 0:
        warnings.append("Colonne date non détectée — vérifiez le format.")
    if cols["label"] < 0:
        warnings.append("Colonne libellé non détectée.")

    transactions: list[dict[str, Any]] = []
    for index in range(header_index + 1, len(rows)):
        row = rows[index]
        if not row or all(not str(cell).strip() for cell in row):
            continue
        txn = row_to_transaction(row, cols, index)
        if txn:
            transactions.append(txn)

    if not transactions:
        warnings.append("Aucun mouvement bancaire détecté dans le fichier.")

    return transactions, warnings


def parse_csv_bytes(content: bytes) -> list[list[Any]]:
    text = content.decode("utf-8-sig", errors="replace")
    delimiter = detect_delimiter(text)
    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    return [list(row) for row in reader]


def parse_xlsx_bytes(content: bytes) -> list[list[Any]]:
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    return [list(row) for row in sheet.iter_rows(values_only=True)]


def parse_bank_file(filename: str, content: bytes) -> tuple[list[dict[str, Any]], dict[str, str], list[str]]:
    lower = (filename or "").lower()
    warnings: list[str] = []

    if lower.endswith((".csv", ".txt")):
        rows = parse_csv_bytes(content)
    elif lower.endswith((".xlsx", ".xls")):
        rows = parse_xlsx_bytes(content)
    else:
        return [], {"filename": filename, "bankName": "BANQUE", "bankIce": "", "bankIf": ""}, [
            "Format tableur non reconnu — utilisez CSV ou Excel."
        ]

    transactions, parse_warnings = parse_bank_rows(rows)
    warnings.extend(parse_warnings)
    meta = {
        "filename": filename,
        "bankName": "BANQUE",
        "bankIce": "",
        "bankIf": "",
    }
    return transactions, meta, warnings


def normalize_ai_transactions(raw_list: list[dict[str, Any]]) -> list[dict[str, Any]]:
    transactions: list[dict[str, Any]] = []
    for index, item in enumerate(raw_list or []):
        label = str(item.get("label") or item.get("libelle") or item.get("description") or "").strip()
        date = parse_date(item.get("date"))
        amount = parse_amount(item.get("amount", item.get("montant")))
        if amount is None:
            continue

        normalized_label = normalize_text(label)
        if SKIP_KEYWORDS.search(normalized_label):
            continue

        const declaredType = str(item.get("type") or "").strip().lower()
        if declaredType in {"payment", "fee"}:
            amount = -abs(amount)
        elif declaredType == "credit":
            amount = abs(amount)

        is_debit = amount < 0
        is_fee = declaredType == "fee" or (FEE_KEYWORDS.search(normalized_label) and is_debit)
        if is_fee:
            txn_type = "fee"
        elif is_debit:
            txn_type = "payment"
        else:
            txn_type = declaredType or "credit"

        abs_amount = abs(amount)
        transactions.append(
            {
                "id": item.get("id") or f"srv-{index}",
                "date": date,
                "label": label,
                "amount": abs_amount * (-1 if is_debit else 1),
                "absAmount": abs_amount,
                "type": txn_type,
            }
        )
    return transactions
